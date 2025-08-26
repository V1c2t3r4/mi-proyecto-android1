import pandas as pd
import re
import unicodedata
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk, Label, Button, Entry, StringVar, filedialog, messagebox, Checkbutton, BooleanVar, DISABLED, NORMAL, scrolledtext, Toplevel
import os
from io import BytesIO

# -------------------- Utilidades --------------------
def clean_sheetname(name: str) -> str:
    return re.sub(r'[\[\]\*\?\/\\:]', '_', str(name))[:31]

def agregar_se(nombre: str) -> str:
    nombre = str(nombre).strip().upper()
    return nombre if nombre.startswith('S/E ') else f'S/E {nombre}'

def norm_txt(s: str) -> str:
    s = str(s).upper()
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
    s = re.sub(r'[^A-Z0-9/ \-]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def cat_estado(v: str) -> str:
    u = str(v).upper()
    if 'CONECT' in u:
        return 'CONECTADO'
    if 'SCR' in u:
        return 'SCR'
    if 'ICC' in u:
        return 'ICC'
    return 'OTROS'

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

# -------------------- Núcleo de cálculo --------------------
def construir_resumen(path_trans, path_info):
    # Leer
    df_trans = pd.read_excel(path_trans)
    df_info  = pd.read_excel(path_info)

    # Transformadores
    df_trans['Nombre Subestación']    = df_trans['Nombre Subestación'].astype(str).str.strip().str.upper()
    df_trans['Nombre Subestación EQ'] = df_trans['Nombre Subestación']
    df_trans['Subestacion_Key']       = df_trans['Nombre Subestación EQ'].map(norm_txt)
    df_trans['Nombre']    = df_trans['Nombre'].astype(str).str.strip()
    df_trans['Capacidad'] = pd.to_numeric(df_trans['Capacidad'], errors='coerce').fillna(0)

    # Info importante
    df_info['SUBESTACION']     = df_info['SUBESTACION'].astype(str).str.strip().str.upper()
    df_info['SUBESTACION_EQ']  = df_info['SUBESTACION'].apply(agregar_se)
    df_info['SUBESTACION_KEY'] = df_info['SUBESTACION_EQ'].map(norm_txt)
    df_info['ALIMENTADOR']     = df_info['ALIMENTADOR'].astype(str).str.strip().str.upper()
    df_info['ALIMENTADOR_KEY'] = df_info['ALIMENTADOR'].map(norm_txt)
    df_info['POTENCIA_MW']     = pd.to_numeric(df_info['POTENCIA_MW'], errors='coerce').fillna(0)
    df_info['ESTADO_PMGD_UP']  = df_info['ESTADO_PMGD'].astype(str).str.upper()
    df_info['CATEGORIA_ESTADO']= df_info['ESTADO_PMGD_UP'].map(cat_estado)

    # Subestaciones de interés
    if 'APLICA' not in df_trans.columns:
        raise KeyError("La columna 'APLICA' no existe en Capacidad_Transformadores.xlsx")
    df_trans['_APLICA_NORM'] = df_trans['APLICA'].astype(str).map(norm_txt)
    mask_si = df_trans['_APLICA_NORM'].isin({'SI','SI.','SÍ','SI (SI)','SI SI','SI/SI'})
    subs_interes_keys = set(df_trans.loc[mask_si, 'Subestacion_Key'].dropna().unique())
    df_trans = df_trans[df_trans['Subestacion_Key'].isin(subs_interes_keys)].copy()

    # Detectar columnas ALIMENTADOR*
    def is_alimentador_col(c: str) -> bool:
        return norm_txt(c).startswith('ALIMENTADOR')
    alimentador_cols = [c for c in df_trans.columns if is_alimentador_col(c)]
    for col in alimentador_cols:
        df_trans[col] = df_trans[col].astype(str).str.strip().str.upper().replace('NAN','')
        df_trans[col + '_KEY'] = df_trans[col].map(norm_txt)

    # Explode trafo-alimentador (sin comodín)
    exploded = []
    for _, row in df_trans.iterrows():
        for col in alimentador_cols:
            al_vis = row[col]; al_key = row[col + '_KEY']
            if al_key and al_key != 'NAN' and al_key != '':
                exploded.append({
                    'Subestacion':      row['Nombre Subestación EQ'],
                    'Subestacion_Key':  row['Subestacion_Key'],
                    'Transformador':    row['Nombre'],
                    'Capacidad_Transf': row['Capacidad'],
                    'Alimentador':      al_vis,
                    'Alimentador_Key':  al_key
                })
    df_exp = pd.DataFrame(exploded)

    # Conjunto de alimentadores por trafo
    if not df_exp.empty:
        tr_feedset = (df_exp.groupby(['Subestacion','Subestacion_Key','Transformador'])
                            .agg({'Capacidad_Transf':'first',
                                  'Alimentador_Key': lambda s: tuple(sorted(set(s))),
                                  'Alimentador':     lambda s: sorted(set(s))})
                            .reset_index()
                            .rename(columns={'Alimentador_Key':'FeedSet_Key',
                                             'Alimentador':'FeedSet_Vis'}))
    else:
        tr_feedset = pd.DataFrame(columns=['Subestacion','Subestacion_Key','Transformador',
                                           'Capacidad_Transf','FeedSet_Key','FeedSet_Vis'])

    # Trafós sin alimentadores → feedset vacío
    trafos_sin_feed = df_trans[~df_trans['Nombre'].isin(tr_feedset['Transformador'])][
        ['Nombre Subestación EQ','Subestacion_Key','Nombre','Capacidad']
    ].rename(columns={'Nombre Subestación EQ':'Subestacion',
                      'Nombre':'Transformador',
                      'Capacidad':'Capacidad_Transf'})
    if not trafos_sin_feed.empty:
        trafos_sin_feed['FeedSet_Key'] = [tuple()] * len(trafos_sin_feed)
        trafos_sin_feed['FeedSet_Vis'] = [[] for _ in range(len(trafos_sin_feed))]
        tr_feedset = pd.concat([tr_feedset, trafos_sin_feed], ignore_index=True)

    # Agrupar por (Subestación, mismo set de alimentadores)
    grp = (tr_feedset.groupby(['Subestacion','Subestacion_Key','FeedSet_Key'], as_index=False)
           .agg({'Capacidad_Transf':'sum',
                 'Transformador':    lambda s: ' + '.join(sorted(map(str, s))),
                 'FeedSet_Vis':      lambda lists: sorted(set(sum(lists, [])))}))

    # PMGD por alimentador (solo subestaciones de interés, y alimentadores existentes en trafos)
    df_valid = df_info[df_info['SUBESTACION_KEY'].isin(subs_interes_keys)].copy()
    if not df_exp.empty:
        pairs = df_exp[['Subestacion_Key','Alimentador_Key']].drop_duplicates()
        df_valid = df_valid.merge(pairs, left_on=['SUBESTACION_KEY','ALIMENTADOR_KEY'],
                                  right_on=['Subestacion_Key','Alimentador_Key'], how='inner')
    else:
        df_valid = df_valid.iloc[0:0].copy()

    pmgd_pairs_cat = (df_valid[df_valid['CATEGORIA_ESTADO'].isin(['CONECTADO','ICC','SCR'])]
                      .groupby(['SUBESTACION_KEY','ALIMENTADOR_KEY','CATEGORIA_ESTADO'], as_index=False)['POTENCIA_MW']
                      .sum())

    if not pmgd_pairs_cat.empty:
        pmgd_by_feed = (pmgd_pairs_cat.pivot(index=['SUBESTACION_KEY','ALIMENTADOR_KEY'],
                                             columns='CATEGORIA_ESTADO', values='POTENCIA_MW')
                        .fillna(0.0).reset_index()
                        .rename(columns={'SUBESTACION_KEY':'Subestacion_Key',
                                         'ALIMENTADOR_KEY':'Alimentador_Key'}))
    else:
        pmgd_by_feed = pd.DataFrame(columns=['Subestacion_Key','Alimentador_Key','CONECTADO','ICC','SCR'])
    for col in ['CONECTADO','ICC','SCR']:
        if col not in pmgd_by_feed.columns:
            pmgd_by_feed[col] = 0.0

    def pmgd_por_feedset(sub_key, feedset):
        if not feedset:
            return {'CONECTADO':0.0,'ICC':0.0,'SCR':0.0}
        bloque = pmgd_by_feed[(pmgd_by_feed['Subestacion_Key'] == sub_key) &
                              (pmgd_by_feed['Alimentador_Key'].isin(feedset))]
        return {'CONECTADO': float(bloque['CONECTADO'].sum()),
                'ICC':       float(bloque['ICC'].sum()),
                'SCR':       float(bloque['SCR'].sum())}

    rows = []
    for _, r in grp.iterrows():
        sub_vis  = r['Subestacion']
        sub_key  = r['Subestacion_Key']
        cap_sum  = float(r['Capacidad_Transf'])
        feedset  = list(r['FeedSet_Key'])
        feeds_vis= r['FeedSet_Vis']

        pmgd = pmgd_por_feedset(sub_key, feedset)
        conect = pmgd['CONECTADO']; icc = pmgd['ICC']; scr = pmgd['SCR']
        total_resta = conect + icc + scr
        disp = cap_sum - total_resta

        rows.append({
            'Subestacion': sub_vis,
            'Transformador': r['Transformador'],
            'Capacidad Transf (MW)': cap_sum,
            'PMGD Conectado (MW)': conect,
            'PMGD ICC (MW)': icc,
            'PMGD SCR (MW)': scr,
            'Total Restado (MW)': total_resta,
            'Capacidad Disponible (MW)': disp,
            'Alimentadores': ', '.join(feeds_vis) if feeds_vis else ''
        })

    df_resumen = pd.DataFrame(rows)
    df_resumen = df_resumen[['Subestacion','Transformador','Capacidad Transf (MW)',
                             'PMGD Conectado (MW)','PMGD ICC (MW)','PMGD SCR (MW)',
                             'Total Restado (MW)','Capacidad Disponible (MW)','Alimentadores']]
    # set de subs con capacidad
    subs_cap = set(df_resumen[df_resumen['Capacidad Disponible (MW)'] > 0]['Subestacion'].unique())
    return df_resumen, df_info, subs_cap

def exportar_resumen(df_resumen, out_path):
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, index=False, sheet_name='RESUMEN')
        ws = writer.sheets['RESUMEN']
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_disp = headers.index('Capacidad Disponible (MW)') + 1
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_rojo  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            val = row[idx_disp-1].value
            try:
                val = float(val) if val is not None else 0.0
            except Exception:
                val = 0.0
            fill = fill_verde if val > 0 else fill_rojo
            for cell in row:
                cell.fill = fill
        autosize_columns(ws)

def exportar_detalle_general(df_info, subs_con_capacidad, out_path):
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sub in sorted(subs_con_capacidad):
            det = df_info[df_info['SUBESTACION_EQ'] == sub].copy()
            det = det[['N_PROCESO','NOMBRE','PROPIETARIO','POTENCIA_MW','COMUNA',
                       'ESTADO_PMGD','ALIMENTADOR','POSTE']].sort_values(['ESTADO_PMGD','ALIMENTADOR'])
            sh = clean_sheetname(sub)
            det.to_excel(writer, index=False, sheet_name=sh)
            autosize_columns(writer.sheets[sh])

def exportar_detalle_por_sub(df_info, sub, out_dir):
    det = df_info[df_info['SUBESTACION_EQ'] == sub].copy()
    det = det[['N_PROCESO','NOMBRE','PROPIETARIO','POTENCIA_MW','COMUNA',
               'ESTADO_PMGD','ALIMENTADOR','POSTE']].sort_values(['ESTADO_PMGD','ALIMENTADOR'])
    safe_sub = re.sub(r'[^A-Za-z0-9_-]+', '_', sub)
    out_path = os.path.join(out_dir, f"Detalle_{safe_sub}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        sh = clean_sheetname(sub)
        det.to_excel(writer, index=False, sheet_name=sh)
        autosize_columns(writer.sheets[sh])
    return out_path

# -------------------- UI (Tkinter) --------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title("Capacidad Subestaciones - Resumen y Detalle")
        root.geometry("740x520")
        root.resizable(False, False)

        self.path_trans = StringVar()
        self.path_info  = StringVar()
        self.estado     = StringVar(value="Seleccione archivos y genere el RESUMEN.")
        self.modo_general = BooleanVar(value=False)
        self.subs_input = StringVar()

        Label(root, text="Capacidad_Transformadores.xlsx").place(x=20, y=20)
        Entry(root, textvariable=self.path_trans, width=70).place(x=20, y=45)
        Button(root, text="Buscar...", command=self.sel_trans).place(x=620, y=42)

        Label(root, text="informacion_importante.xlsx").place(x=20, y=85)
        Entry(root, textvariable=self.path_info, width=70).place(x=20, y=110)
        Button(root, text="Buscar...", command=self.sel_info).place(x=620, y=107)

        Button(root, text="Generar SOLO RESUMEN", command=self.run_resumen, width=25).place(x=20, y=155)

        Label(root, text="Detalle (opcional):").place(x=20, y=200)
        Checkbutton(root, text="GENERAL (todas con capacidad)", variable=self.modo_general,
                    command=self.toggle_detalle_mode).place(x=20, y=225)

        Label(root, text="o lista de 1 a 10 subestaciones (separadas por ';')").place(x=20, y=255)
        Entry(root, textvariable=self.subs_input, width=80).place(x=20, y=280)

        Button(root, text="Generar DETALLES", command=self.run_detalles, width=25).place(x=20, y=320)
        Button(root, text="Ver subestaciones con capacidad", command=self.ver_subs_cap).place(x=250, y=320)

        Label(root, text="Estado:").place(x=20, y=365)
        self.log = scrolledtext.ScrolledText(root, width=88, height=7, state='disabled')
        self.log.place(x=20, y=385)

        self.df_resumen = None
        self.df_info = None
        self.subs_con_capacidad = set()
        self.resumen_file = None

    def sel_trans(self):
        p = filedialog.askopenfilename(title="Selecciona Capacidad_Transformadores.xlsx",
                                       filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.path_trans.set(p)

    def sel_info(self):
        p = filedialog.askopenfilename(title="Selecciona informacion_importante.xlsx",
                                       filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.path_info.set(p)

    def log_line(self, txt):
        self.log.configure(state='normal')
        self.log.insert('end', txt + "\n")
        self.log.configure(state='disabled')
        self.log.see('end')

    def toggle_detalle_mode(self):
        # Si es general, bloqueamos la entrada manual para que no se confunda
        if self.modo_general.get():
            self.subs_input.set("")
        # no hace falta deshabilitar la caja; solo limpiar

    def run_resumen(self):
        try:
            if not self.path_trans.get() or not self.path_info.get():
                messagebox.showwarning("Faltan archivos", "Selecciona ambos archivos Excel.")
                return
            self.log_line("Procesando RESUMEN...")
            df_resumen, df_info, subs_cap = construir_resumen(self.path_trans.get(), self.path_info.get())
            self.df_resumen = df_resumen
            self.df_info = df_info
            self.subs_con_capacidad = subs_cap

            out_name = f"Resumen_Subestaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = os.path.join(os.getcwd(), out_name)
            exportar_resumen(df_resumen, out_path)
            self.resumen_file = out_path

            self.log_line(f"✅ RESUMEN creado: {out_path}")
            if subs_cap:
                self.log_line(f"Subestaciones con capacidad: {len(subs_cap)} (usa el botón para ver la lista)")
            else:
                self.log_line("No hay subestaciones con capacidad disponible.")
            messagebox.showinfo("Listo", "RESUMEN generado.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_line(f"Error: {e}")

    def ver_subs_cap(self):
        if not self.df_resumen is not None:
            messagebox.showinfo("Aviso", "Primero genera el RESUMEN.")
            return
        subs = sorted(self.subs_con_capacidad)
        if not subs:
            messagebox.showinfo("Aviso", "No hay subestaciones con capacidad.")
            return
        top = Toplevel(self.root); top.title("Subestaciones con capacidad")
        txt = scrolledtext.ScrolledText(top, width=60, height=20)
        txt.pack(padx=10, pady=10)
        for s in subs:
            txt.insert('end', s + "\n")
        txt.configure(state='disabled')

    def run_detalles(self):
        try:
            if self.df_resumen is None or self.df_info is None:
                messagebox.showwarning("Falta RESUMEN", "Primero genera el RESUMEN.")
                return
            if not self.subs_con_capacidad:
                messagebox.showinfo("Sin capacidad", "No hay subestaciones con capacidad. No se generan detalles.")
                return

            if self.modo_general.get():
                out_name = f"Detalle_Subestaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                out_path = os.path.join(os.getcwd(), out_name)
                exportar_detalle_general(self.df_info, self.subs_con_capacidad, out_path)
                self.log_line(f"📄 Detalle GENERAL creado: {out_path}")
                messagebox.showinfo("Listo", "Detalle GENERAL generado.")
            else:
                txt = self.subs_input.get().strip()
                if not txt:
                    messagebox.showwarning("Falta lista", "Escribe entre 1 y 10 subestaciones separadas por ';' o elige GENERAL.")
                    return
                partes = [p.strip() for p in txt.split(';') if p.strip()]
                if len(partes) == 0 or len(partes) > 10:
                    messagebox.showwarning("Cantidad inválida", "Ingresa entre 1 y 10 subestaciones separadas por ';'.")
                    return

                mapa_vis_to_key = {s: norm_txt(s) for s in self.subs_con_capacidad}
                generados = 0
                out_dir = os.getcwd()
                for entrada in partes:
                    key_in = norm_txt(agregar_se(entrada))
                    candidatos = [vis for vis, k in mapa_vis_to_key.items() if k == key_in]
                    if not candidatos:
                        self.log_line(f"❌ No coincide o no tiene capacidad: {entrada}")
                        continue
                    sub_vis = candidatos[0]
                    out_path = exportar_detalle_por_sub(self.df_info, sub_vis, out_dir)
                    self.log_line(f"📄 Detalle creado: {out_path}")
                    generados += 1
                if generados == 0:
                    messagebox.showinfo("Sin detalles", "No se generó ningún detalle (revisa nombres y capacidad).")
                else:
                    messagebox.showinfo("Listo", f"Detalles generados: {generados}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_line(f"Error: {e}")

# -------------------- Main --------------------
if __name__ == "__main__":
    root = Tk()
    App(root)
    root.mainloop()
